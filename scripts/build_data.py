from __future__ import annotations

import argparse
import csv
import json
import os
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = ROOT / "public" / "data" / "dashboard-data.json"

BOOKING_CANDIDATES = [
    "Booking_Productwise_Report (3).csv",
    "Booking_Productwise_Report.csv",
]
HIERARCHY_CANDIDATES = [
    "Hierarchy data latest 30.03.2026.xlsx",
    "Hierarchy.xlsx",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build sanitized AP Circle dashboard JSON from bookings CSV and hierarchy XLSX."
    )
    parser.add_argument("--bookings", help="Path to Booking_Productwise_Report CSV.")
    parser.add_argument("--hierarchy", help="Path to hierarchy XLSX workbook.")
    parser.add_argument("--out", default=str(OUT_PATH), help="Output JSON path.")
    return parser.parse_args()


def resolve_source(explicit: str | None, env_name: str, candidates: list[str]) -> Path:
    if explicit:
        path = Path(explicit)
        if path.exists():
            return path
        raise FileNotFoundError(f"Source path not found: {path}")

    env_value = os.environ.get(env_name)
    if env_value:
        path = Path(env_value)
        if path.exists():
            return path
        raise FileNotFoundError(f"{env_name} path not found: {path}")

    search_dirs = [ROOT / "data" / "raw", Path.home() / "Downloads"]
    for directory in search_dirs:
        for candidate in candidates:
            path = directory / candidate
            if path.exists():
                return path

    searched = ", ".join(str(d / c) for d in search_dirs for c in candidates)
    raise FileNotFoundError(f"Could not find source file. Searched: {searched}")


def text(value: Any, fallback: str = "") -> str:
    if value is None:
        return fallback
    clean = str(value).strip()
    return clean if clean else fallback


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(str(value).replace(",", "").strip())


def integer(value: Any) -> int:
    return int(number(value))


def round2(value: float) -> float:
    return round(value + 0.0000001, 2)


def product_bucket(product_name: str) -> str:
    lowered = product_name.lower()
    if "speed post" in lowered:
        return "speedPost"
    if "parcel" in lowered:
        return "parcels"
    return "other"


def office_category(office_type_code: str) -> str:
    if office_type_code == "HPO":
        return "HO"
    if office_type_code == "SPO":
        return "SO"
    if office_type_code == "BPO":
        return "BO"
    return "Others"


def division_group(division_name: str, region_name: str) -> str:
    if not division_name or division_name == "Unmapped / Missing hierarchy":
        return "missing"
    if "RMS" in division_name.upper():
        return "rms"
    if division_name == region_name or division_name.endswith(" Region"):
        return "adminOther"
    return "postal"


def empty_metrics() -> dict[str, Any]:
    return {
        "transactions": 0,
        "revenue": 0.0,
        "rowCount": 0,
        "bucketTransactions": {"speedPost": 0, "parcels": 0, "other": 0},
        "bucketRevenue": {"speedPost": 0.0, "parcels": 0.0, "other": 0.0},
        "speedPostDetails": defaultdict(lambda: {"transactions": 0, "revenue": 0.0}),
    }


def build_hierarchy(hierarchy_path: Path) -> dict[int, dict[str, Any]]:
    workbook = load_workbook(hierarchy_path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = list(next(worksheet.iter_rows(values_only=True)))
    index = {name: position for position, name in enumerate(headers)}

    required = [
        "office_id",
        "office_name",
        "office_type_code",
        "office_type_desc",
        "office_status",
        "region_name",
        "division_name",
        "sub_division_name",
        "ho_name",
        "so_name",
        "bo_name",
    ]
    missing = [column for column in required if column not in index]
    if missing:
        raise ValueError(f"Hierarchy workbook is missing required columns: {missing}")

    offices: dict[int, dict[str, Any]] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        office_id = row[index["office_id"]]
        if office_id is None:
            continue

        office_type = text(row[index["office_type_code"]])
        region = text(row[index["region_name"]], "Unmapped")
        division = text(row[index["division_name"]], "Unmapped")
        offices[int(office_id)] = {
            "officeId": int(office_id),
            "officeName": text(row[index["office_name"]], f"Office {office_id}"),
            "officeTypeCode": office_type,
            "officeTypeDesc": text(row[index["office_type_desc"]], office_type),
            "officeStatus": text(row[index["office_status"]], "Unknown"),
            "regionName": region,
            "divisionName": division,
            "divisionGroup": division_group(division, region),
            "subDivisionName": text(row[index["sub_division_name"]], "Unmapped"),
            "hoName": text(row[index["ho_name"]]),
            "soName": text(row[index["so_name"]]),
            "boName": text(row[index["bo_name"]]),
            "category": office_category(office_type),
            "missingHierarchy": False,
        }
    return offices


def read_bookings(bookings_path: Path) -> tuple[dict[int, dict[str, Any]], dict[str, Any]]:
    metrics: dict[int, dict[str, Any]] = defaultdict(empty_metrics)
    product_totals: dict[str, dict[str, float]] = defaultdict(
        lambda: {"transactions": 0, "revenue": 0.0}
    )
    bucket_totals = {
        "speedPost": {"transactions": 0, "revenue": 0.0},
        "parcels": {"transactions": 0, "revenue": 0.0},
        "other": {"transactions": 0, "revenue": 0.0},
    }
    booking_office_names: dict[int, str] = {}
    booking_dates: list[str] = []
    negative_row_examples: list[dict[str, Any]] = []
    negative_row_count = 0
    csv_rows = 0

    with bookings_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        required = [
            "office-id",
            "office-name",
            "booking-date",
            "product-name",
            "article-count",
            "tax",
            "total_amount",
        ]
        missing = [column for column in required if column not in (reader.fieldnames or [])]
        if missing:
            raise ValueError(f"Bookings CSV is missing required columns: {missing}")

        for row in reader:
            csv_rows += 1
            office_id = integer(row["office-id"])
            office_name = text(row["office-name"], f"Office {office_id}")
            product_name = text(row["product-name"], "Unknown product")
            transactions = integer(row["article-count"])
            tax = number(row["tax"])
            total_amount = number(row["total_amount"])
            revenue = total_amount - tax
            bucket = product_bucket(product_name)
            date_value = text(row["booking-date"])[:10]

            if date_value:
                booking_dates.append(date_value)
            if revenue < 0:
                negative_row_count += 1
                if len(negative_row_examples) < 25:
                    negative_row_examples.append(
                        {
                            "officeId": office_id,
                            "officeName": office_name,
                            "productName": product_name,
                            "transactions": transactions,
                            "tax": round2(tax),
                            "totalAmount": round2(total_amount),
                            "revenue": round2(revenue),
                        }
                    )

            booking_office_names[office_id] = office_name
            office_metrics = metrics[office_id]
            office_metrics["transactions"] += transactions
            office_metrics["revenue"] += revenue
            office_metrics["rowCount"] += 1
            office_metrics["bucketTransactions"][bucket] += transactions
            office_metrics["bucketRevenue"][bucket] += revenue

            if bucket == "speedPost":
                detail = office_metrics["speedPostDetails"][product_name]
                detail["transactions"] += transactions
                detail["revenue"] += revenue

            product_totals[product_name]["transactions"] += transactions
            product_totals[product_name]["revenue"] += revenue
            bucket_totals[bucket]["transactions"] += transactions
            bucket_totals[bucket]["revenue"] += revenue

    metadata = {
        "csvRows": csv_rows,
        "uniqueBookingOffices": len(booking_office_names),
        "bookingOfficeNames": booking_office_names,
        "dateStart": min(booking_dates) if booking_dates else "",
        "dateEnd": max(booking_dates) if booking_dates else "",
        "negativeRevenueRowCount": negative_row_count,
        "negativeRevenueRowExamples": negative_row_examples,
        "productTotals": [
            {
                "productName": product_name,
                "transactions": int(values["transactions"]),
                "revenue": round2(values["revenue"]),
                "bucket": product_bucket(product_name),
            }
            for product_name, values in sorted(product_totals.items())
        ],
        "bucketTotals": {
            key: {
                "transactions": int(values["transactions"]),
                "revenue": round2(values["revenue"]),
            }
            for key, values in bucket_totals.items()
        },
    }
    return metrics, metadata


def finalize_offices(
    hierarchy: dict[int, dict[str, Any]],
    booking_metrics: dict[int, dict[str, Any]],
    booking_office_names: dict[int, str],
) -> list[dict[str, Any]]:
    office_ids = sorted(set(hierarchy) | set(booking_metrics))
    offices: list[dict[str, Any]] = []

    for office_id in office_ids:
        base = hierarchy.get(office_id)
        if base is None:
            office_name = booking_office_names.get(office_id, f"Office {office_id}")
            base = {
                "officeId": office_id,
                "officeName": office_name,
                "officeTypeCode": "MISSING",
                "officeTypeDesc": "Missing hierarchy",
                "officeStatus": "Unknown",
                "regionName": "Unmapped",
                "divisionName": "Unmapped / Missing hierarchy",
                "divisionGroup": "missing",
                "subDivisionName": "Unmapped",
                "hoName": "",
                "soName": "",
                "boName": "",
                "category": "Others",
                "missingHierarchy": True,
            }

        metrics = booking_metrics.get(office_id, empty_metrics())
        transactions = int(metrics["transactions"])
        revenue = round2(metrics["revenue"])
        target_band = "Not BO"
        if base["category"] == "BO" and base["officeStatus"] == "Active":
            if transactions == 0:
                target_band = "Nil"
            elif transactions <= 10:
                target_band = "1-10"
            else:
                target_band = ">10"

        speed_details = [
            {
                "productName": product_name,
                "transactions": int(values["transactions"]),
                "revenue": round2(values["revenue"]),
            }
            for product_name, values in sorted(
                metrics["speedPostDetails"].items(),
                key=lambda item: (-item[1]["transactions"], item[0]),
            )
        ]

        office = {
            **base,
            "transactions": transactions,
            "revenue": revenue,
            "rowCount": int(metrics["rowCount"]),
            "bucketTransactions": {
                key: int(value) for key, value in metrics["bucketTransactions"].items()
            },
            "bucketRevenue": {
                key: round2(value) for key, value in metrics["bucketRevenue"].items()
            },
            "speedPostDetails": speed_details,
            "targetBand": target_band,
            "negativeRevenue": revenue < 0,
        }
        offices.append(office)
    return offices


def summarize_group(offices: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for office in offices:
        grouped[office[key]].append(office)

    summaries = []
    for name, group_offices in grouped.items():
        active_bos = [
            office
            for office in group_offices
            if office["category"] == "BO" and office["officeStatus"] == "Active"
        ]
        bucket_transactions = Counter()
        bucket_revenue = Counter()
        category_counts = Counter()
        for office in group_offices:
            category_counts[office["category"]] += 1
            for bucket, value in office["bucketTransactions"].items():
                bucket_transactions[bucket] += value
            for bucket, value in office["bucketRevenue"].items():
                bucket_revenue[bucket] += value

        summaries.append(
            {
                "name": name,
                "regionName": group_offices[0]["regionName"],
                "divisionGroup": group_offices[0]["divisionGroup"],
                "officeCount": len(group_offices),
                "officeCountsByCategory": {
                    "HO": category_counts["HO"],
                    "SO": category_counts["SO"],
                    "BO": category_counts["BO"],
                    "Others": category_counts["Others"],
                },
                "activeBOs": len(active_bos),
                "nilBOs": sum(1 for office in active_bos if office["targetBand"] == "Nil"),
                "lowBOs": sum(1 for office in active_bos if office["targetBand"] == "1-10"),
                "aboveTargetBOs": sum(
                    1 for office in active_bos if office["targetBand"] == ">10"
                ),
                "transactions": sum(office["transactions"] for office in group_offices),
                "revenue": round2(sum(office["revenue"] for office in group_offices)),
                "bucketTransactions": {
                    "speedPost": int(bucket_transactions["speedPost"]),
                    "parcels": int(bucket_transactions["parcels"]),
                    "other": int(bucket_transactions["other"]),
                },
                "bucketRevenue": {
                    "speedPost": round2(bucket_revenue["speedPost"]),
                    "parcels": round2(bucket_revenue["parcels"]),
                    "other": round2(bucket_revenue["other"]),
                },
            }
        )

    return sorted(summaries, key=lambda item: item["name"])


def build_payload(bookings_path: Path, hierarchy_path: Path) -> dict[str, Any]:
    hierarchy = build_hierarchy(hierarchy_path)
    booking_metrics, booking_meta = read_bookings(bookings_path)
    booking_office_names = booking_meta.pop("bookingOfficeNames")
    offices = finalize_offices(hierarchy, booking_metrics, booking_office_names)

    active_bos = [
        office
        for office in offices
        if office["category"] == "BO" and office["officeStatus"] == "Active"
    ]
    missing_hierarchy = [office for office in offices if office["missingHierarchy"]]
    negative_offices = [office for office in offices if office["negativeRevenue"]]
    named_divisions = {
        office["divisionName"]: office["divisionGroup"]
        for office in offices
        if not office["missingHierarchy"] and office["divisionName"] != "Unmapped"
    }
    division_counts = Counter(named_divisions.values())

    bucket_transactions = Counter()
    bucket_revenue = Counter()
    for office in offices:
        for bucket, value in office["bucketTransactions"].items():
            bucket_transactions[bucket] += value
        for bucket, value in office["bucketRevenue"].items():
            bucket_revenue[bucket] += value

    total_transactions = sum(office["transactions"] for office in offices)
    total_revenue = sum(office["revenue"] for office in offices)
    data_quality = {
        "nilBOs": [
            {
                "officeId": office["officeId"],
                "officeName": office["officeName"],
                "regionName": office["regionName"],
                "divisionName": office["divisionName"],
            }
            for office in active_bos
            if office["targetBand"] == "Nil"
        ],
        "missingHierarchyOffices": [
            {
                "officeId": office["officeId"],
                "officeName": office["officeName"],
                "transactions": office["transactions"],
                "revenue": office["revenue"],
            }
            for office in missing_hierarchy
        ],
        "negativeRevenueOffices": [
            {
                "officeId": office["officeId"],
                "officeName": office["officeName"],
                "regionName": office["regionName"],
                "divisionName": office["divisionName"],
                "revenue": office["revenue"],
            }
            for office in negative_offices
        ],
        "negativeRevenueRowCount": booking_meta["negativeRevenueRowCount"],
        "negativeRevenueRowExamples": booking_meta["negativeRevenueRowExamples"],
    }

    return {
        "metadata": {
            "sourceFiles": {
                "bookings": bookings_path.name,
                "hierarchy": hierarchy_path.name,
            },
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "dateStart": booking_meta["dateStart"],
            "dateEnd": booking_meta["dateEnd"],
            "rowCounts": {
                "csvRows": booking_meta["csvRows"],
                "uniqueBookingOffices": booking_meta["uniqueBookingOffices"],
                "hierarchyOffices": len(hierarchy),
                "generatedOffices": len(offices),
            },
            "rules": {
                "transactionRule": "Transactions use article-count summed across the selected report period.",
                "revenueRule": "Revenue earned = total_amount - tax. Negative values are preserved and flagged.",
                "targetRule": "Active BO target band across the report period: Nil, 1-10, >10.",
                "productBuckets": {
                    "speedPost": "Product name contains 'Speed Post', including Speed Post Parcel Domestic.",
                    "parcels": "Product name contains 'parcel' after excluding Speed Post products.",
                    "other": "All remaining booking products.",
                },
            },
        },
        "circle": {
            "name": "Andhra Pradesh Circle",
            "activeBOs": len(active_bos),
            "nilBOs": sum(1 for office in active_bos if office["targetBand"] == "Nil"),
            "lowBOs": sum(1 for office in active_bos if office["targetBand"] == "1-10"),
            "aboveTargetBOs": sum(
                1 for office in active_bos if office["targetBand"] == ">10"
            ),
            "transactions": int(total_transactions),
            "revenue": round2(total_revenue),
            "bucketTransactions": {
                "speedPost": int(bucket_transactions["speedPost"]),
                "parcels": int(bucket_transactions["parcels"]),
                "other": int(bucket_transactions["other"]),
            },
            "bucketRevenue": {
                "speedPost": round2(bucket_revenue["speedPost"]),
                "parcels": round2(bucket_revenue["parcels"]),
                "other": round2(bucket_revenue["other"]),
            },
            "divisionCounts": {
                "postal": division_counts["postal"],
                "rms": division_counts["rms"],
                "adminOther": division_counts["adminOther"],
            },
        },
        "regions": summarize_group(
            [office for office in offices if office["regionName"] != "Unmapped"],
            "regionName",
        ),
        "divisions": summarize_group(
            [
                office
                for office in offices
                if office["divisionGroup"] in {"postal", "rms", "adminOther"}
            ],
            "divisionName",
        ),
        "offices": offices,
        "productTotals": booking_meta["productTotals"],
        "dataQuality": data_quality,
    }


def main() -> None:
    args = parse_args()
    bookings_path = resolve_source(args.bookings, "BOOKING_CSV", BOOKING_CANDIDATES)
    hierarchy_path = resolve_source(args.hierarchy, "HIERARCHY_XLSX", HIERARCHY_CANDIDATES)
    payload = build_payload(bookings_path, hierarchy_path)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print(f"Wrote {out_path}")
    print(
        "Rows: {csvRows:,} | Offices: {generatedOffices:,} | Active BOs: {activeBOs:,}".format(
            csvRows=payload["metadata"]["rowCounts"]["csvRows"],
            generatedOffices=payload["metadata"]["rowCounts"]["generatedOffices"],
            activeBOs=payload["circle"]["activeBOs"],
        )
    )


if __name__ == "__main__":
    main()
